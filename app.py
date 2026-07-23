"""
Ampyr Solar Europe — FAC & 1st Year Testing Dashboard
Light / pastel Streamlit dashboard built on the FAC Tracker (Gantt) workbook.
Leads with real completion status (from the Status column) and keeps date
timing as a secondary view.

Run:  streamlit run app.py
"""

from __future__ import annotations

import datetime as dt
import glob as _glob
import io
import os
import openpyxl
from datetime import date, datetime, timedelta

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

# ===================================================================
# Embedded parser (merged from fac_parser.py so the app is a single
# self-contained file — only app.py + the Excel need to be deployed).
# ===================================================================
from datetime import date, datetime, timedelta

import openpyxl
import pandas as pd

CATEGORY_COLORS = {
    "FFC4B5FD": "Prerequisite",
    "FF6EE7B7": "Compliance",
    "FFFDE68A": "Closure",
    "FFFED7AA": "Financial",
    "FFA5F3FC": "Technical",
    "FFFBCFE8": "Testing",
    "FFBFDBFE": "Document",
}

TODAY_MARKER_FILL = "FFFEE2E2"
DEFAULT_ANCHOR_DATE = date(2026, 7, 14)
FIRST_METADATA_ROW = 4

# metadata column positions (1-based) — stable, Status/Gantt are added after these
COL_SITE, COL_FACDATE, COL_OM = 1, 2, 3
COL_TASK, COL_SUBTASK, COL_RESP, COL_ACCEPT, COL_TESTDATE = 4, 5, 6, 7, 8

DONE_TOKENS = {"done", "yes", "y", "complete", "completed", "✓", "true", "x"}


def _fill_hex(cell):
    fill = cell.fill
    if fill and fill.patternType:
        rgb = fill.fgColor.rgb
        if isinstance(rgb, str):
            return rgb
    return None


def _detect_layout(ws):
    """Return (status_col or None, first_gantt_col)."""
    status_col = None
    gantt_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str) and v.strip().lower() == "status":
            status_col = c
        if gantt_col is None and v is not None:
            s = str(v).strip()
            if s.isdigit() and len(s) == 4 and 2000 <= int(s) <= 2100:
                gantt_col = c
    if gantt_col is None:
        gantt_col = (status_col + 1) if status_col else 9
    return status_col, gantt_col


def _find_anchor(ws, first_gantt_col):
    for r in range(1, 4):
        for c in range(first_gantt_col, ws.max_column + 1):
            if _fill_hex(ws.cell(row=r, column=c)) == TODAY_MARKER_FILL:
                return c, DEFAULT_ANCHOR_DATE
    return first_gantt_col + 80, DEFAULT_ANCHOR_DATE


def _col_to_date(col, anchor_col, anchor_date):
    return anchor_date + timedelta(weeks=(col - anchor_col))


def _is_done(v):
    return isinstance(v, str) and v.strip().lower() in DONE_TOKENS or v is True


def _parse_header(text):
    parts = [p.strip() for p in str(text).split("|")]
    name = parts[0].strip() if parts else str(text).strip()
    country, fac_raw, contractor = "", "", ""
    for p in parts[1:]:
        low = p.lower()
        if low.startswith("fac"):
            fac_raw = p.split(":", 1)[1].strip() if ":" in p else ""
        elif low.startswith("o&m") or low.startswith("o & m"):
            contractor = p.split(":", 1)[1].strip() if ":" in p else ""
        elif p and not country:
            country = p
    return name, country, contractor, fac_raw


def _parse_date_any(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    s = str(v).strip()
    if not s or s in {"—", "-", "TBC", "TBD"}:
        return None
    for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _timing(d, today):
    if d is None:
        return "TBC"
    delta = (d - today).days
    if delta < 0:
        return "Date passed"
    if delta <= 90:
        return "Due soon"
    return "Scheduled"


def _completion_status(done, total):
    if total == 0:
        return "No tasks"
    if done >= total:
        return "Completed"
    if done > 0:
        return "In progress"
    return "Not started"


def parse_workbook(source, today=None):
    if today is None:
        today = date.today()
    if isinstance(today, datetime):
        today = today.date()

    wb = openpyxl.load_workbook(source, data_only=True)
    ws = wb[wb.sheetnames[0]]
    status_col, first_gantt = _detect_layout(ws)
    anchor_col, anchor_date = _find_anchor(ws, first_gantt)

    sites, milestones, tasks = [], [], []
    cur_site = cur_phase = cur_task = None
    phase_agg = {}  # (site, phase) -> [done, total]

    for r in range(FIRST_METADATA_ROW, ws.max_row + 1):
        A = ws.cell(row=r, column=COL_SITE).value
        D = ws.cell(row=r, column=COL_TASK).value
        E = ws.cell(row=r, column=COL_SUBTASK).value
        F = ws.cell(row=r, column=COL_RESP).value
        G = ws.cell(row=r, column=COL_ACCEPT).value
        H = ws.cell(row=r, column=COL_TESTDATE).value
        S = ws.cell(row=r, column=status_col).value if status_col else None

        a_txt = str(A).strip() if A is not None else ""
        d_txt = str(D).strip() if D is not None else ""
        e_txt = str(E).strip() if E is not None else ""
        done_flag = _is_done(S)

        if a_txt and "|" in a_txt:
            name, country, contractor, fac_raw = _parse_header(a_txt)
            cur_site = {"Site": name, "Country": country, "Contractor": contractor,
                        "FAC Date": _parse_date_any(fac_raw), "FYT Date": None}
            sites.append(cur_site)
            cur_phase = cur_task = None
            continue

        if cur_site is None:
            continue
        if a_txt and a_txt.lower() == cur_site["Site"].lower():
            continue

        if d_txt.startswith("▸") or d_txt.startswith(">"):
            up = d_txt.upper()
            if "FIRST YEAR" in up:
                cur_phase = "First Year Testing"
            elif "FINAL ACCEPTANCE" in up:
                cur_phase = "FAC"
            else:
                cur_phase = d_txt.lstrip("▸> ").strip()
            phase_date = _parse_date_any(H)
            if cur_phase == "First Year Testing":
                cur_site["FYT Date"] = phase_date
            milestones.append({"Site": cur_site["Site"], "Country": cur_site["Country"],
                               "Contractor": cur_site["Contractor"],
                               "Milestone": cur_phase, "Date": phase_date,
                               "PhaseHeaderDone": done_flag})
            phase_agg.setdefault((cur_site["Site"], cur_phase), [0, 0])
            cur_task = None
            continue

        if d_txt:
            cur_task = d_txt
        if not (d_txt or e_txt):
            continue

        start_col = end_col = None
        counts = {}
        for c in range(first_gantt, ws.max_column + 1):
            hexc = _fill_hex(ws.cell(row=r, column=c))
            if hexc in CATEGORY_COLORS:
                if start_col is None:
                    start_col = c
                end_col = c
                cat = CATEGORY_COLORS[hexc]
                counts[cat] = counts.get(cat, 0) + 1
        category = max(counts, key=counts.get) if counts else ""
        start_date = _col_to_date(start_col, anchor_col, anchor_date) if start_col else None
        end_date = (_col_to_date(end_col, anchor_col, anchor_date) + timedelta(days=6)) if end_col else None

        key = (cur_site["Site"], cur_phase)
        agg = phase_agg.setdefault(key, [0, 0])
        agg[1] += 1
        if done_flag:
            agg[0] += 1

        tasks.append({"Site": cur_site["Site"], "Country": cur_site["Country"],
                      "Contractor": cur_site["Contractor"], "Phase": cur_phase or "",
                      "Task": cur_task or "", "Sub-Task": e_txt,
                      "Responsibility": str(F).strip() if F else "",
                      "Accepted By": str(G).strip() if G else "",
                      "Category": category, "Start": start_date, "End": end_date,
                      "Status": "Done" if done_flag else "Pending"})

    df_sites = pd.DataFrame(sites)
    df_ms = pd.DataFrame(milestones)
    df_tasks = pd.DataFrame(tasks)

    # ---- roll up completion onto milestones & sites ----
    def phase_done_total(site, phase):
        return phase_agg.get((site, phase), [0, 0])

    if not df_ms.empty:
        st_list, done_list, tot_list, pct_list, timing_list = [], [], [], [], []
        for _, r in df_ms.iterrows():
            d, t = phase_done_total(r["Site"], r["Milestone"])
            if t > 0 and d == t:
                header_done = True
            else:
                header_done = bool(r.get("PhaseHeaderDone"))
                if header_done:  # phase flagged done at header -> treat as complete
                    d, t = (t, t) if t else (1, 1)
            status = _completion_status(d, t)
            done_list.append(d); tot_list.append(t)
            pct_list.append(round(100 * d / t) if t else 0)
            st_list.append(status)
            timing_list.append(_timing(r["Date"], today))
        df_ms["Done"] = done_list
        df_ms["Total"] = tot_list
        df_ms["Progress %"] = pct_list
        df_ms["Status"] = st_list
        df_ms["Timing"] = timing_list

    if not df_sites.empty:
        for lab, phase in [("FYT", "First Year Testing"), ("FAC", "FAC")]:
            dts = df_sites["Site"].apply(lambda s: phase_done_total(s, phase))
            df_sites[f"{lab} Done"] = dts.apply(lambda x: x[0])
            df_sites[f"{lab} Total"] = dts.apply(lambda x: x[1])
            df_sites[f"{lab} Progress %"] = dts.apply(
                lambda x: round(100 * x[0] / x[1]) if x[1] else 0)
            df_sites[f"{lab} Status"] = dts.apply(lambda x: _completion_status(x[0], x[1]))
        # honour a phase-header "Done" flag even when task rows are blank
        if not df_ms.empty:
            for lab, phase in [("FYT", "First Year Testing"), ("FAC", "FAC")]:
                hdr = df_ms[df_ms["Milestone"] == phase].set_index("Site")["Status"]
                df_sites[f"{lab} Status"] = df_sites.apply(
                    lambda row: hdr.get(row["Site"], row[f"{lab} Status"]), axis=1)
        df_sites["FAC Timing"] = df_sites["FAC Date"].apply(lambda d: _timing(d, today))
        df_sites["FYT Timing"] = df_sites["FYT Date"].apply(lambda d: _timing(d, today))

    return {"sites": df_sites, "milestones": df_ms, "tasks": df_tasks,
            "today": today, "anchor_col": anchor_col, "anchor_date": anchor_date,
            "has_status": status_col is not None}
# ---- end embedded parser ----


_HERE = os.path.dirname(os.path.abspath(__file__))


def _find_default_workbook():
    candidates = [
        os.path.join(_HERE, "data", "FAC_Tracker_Gantt.xlsx"),
        os.path.join(_HERE, "FAC_Tracker_Gantt.xlsx"),
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    for pattern in (os.path.join(_HERE, "data", "*.xlsx"),
                    os.path.join(_HERE, "*.xlsx")):
        hits = [h for h in _glob.glob(pattern)
                if not os.path.basename(h).startswith("~$")]
        if hits:
            return sorted(hits)[0]
    return None


DATA_DEFAULT = _find_default_workbook()

# ---- palettes ---------------------------------------------------------------
PASTEL = {
    "Prerequisite": "#C4B5FD", "Compliance": "#6EE7B7", "Closure": "#FDE68A",
    "Financial": "#FED7AA", "Technical": "#A5F3FC", "Testing": "#FBCFE8",
    "Document": "#BFDBFE",
}
# completion status (headline)
COMP_COLORS = {
    "Completed": "#A7F3D0", "In progress": "#FDE68A",
    "Not started": "#FBC7CE", "No tasks": "#E5E7EB", "TBC": "#E5E7EB",
}
COMP_TEXT = {
    "Completed": "#065F46", "In progress": "#92400E",
    "Not started": "#9F1239", "No tasks": "#374151", "TBC": "#374151",
}
# date timing (secondary)
TIME_COLORS = {
    "Date passed": "#FCA5A5", "Due soon": "#FDE68A",
    "Scheduled": "#A7F3D0", "TBC": "#E5E7EB",
}
TIME_TEXT = {
    "Date passed": "#991B1B", "Due soon": "#92400E",
    "Scheduled": "#065F46", "TBC": "#374151",
}
TASK_COLORS = {"Done": "#A7F3D0", "Pending": "#E5E7EB"}
GREEN_DONE = "#34D399"          # milestone dot when completed
FYT_COL_BG = "#EAF2FF"          # 1st Year Testing column tint (blue)
FAC_COL_BG = "#FFF1E6"          # FAC column tint (peach)
TASK_TEXT = {"Done": "#065F46", "Pending": "#4b5563"}

st.set_page_config(page_title="FAC & 1st Year Tracking",
                   page_icon="🌤️", layout="wide")

st.markdown("""
<style>
:root { --card-radius: 16px; }
.stApp { background: linear-gradient(160deg,#fbfcff 0%,#f4f7fb 45%,#fef6fb 100%); }
section[data-testid="stSidebar"] { background: #f7f5ff; }
h1,h2,h3 { color:#3f3d56; font-weight:700; }
.block-container { padding-top: 1.6rem; }
.kpi { background:#ffffff; border-radius:var(--card-radius); padding:18px 20px;
  box-shadow:0 3px 14px rgba(120,120,160,0.10); border:1px solid #eef0f7; height:100%; }
.kpi .label { font-size:0.82rem; color:#7c7a94; font-weight:600;
  text-transform:uppercase; letter-spacing:.4px; }
.kpi .value { font-size:2.0rem; font-weight:800; color:#3f3d56; line-height:1.1; }
.kpi .sub { font-size:0.8rem; color:#9a98b0; margin-top:2px; }
.pill { padding:3px 12px; border-radius:999px; font-size:0.78rem; font-weight:700;
  display:inline-block; }
.legend-chip { display:inline-block; padding:3px 10px; margin:2px 4px 2px 0;
  border-radius:999px; font-size:0.75rem; color:#3f3d56; font-weight:600; }
.bar-wrap { background:#eef0f7; border-radius:999px; height:14px; width:100%; overflow:hidden; }
.bar-fill { height:14px; border-radius:999px; }
.site-row { padding:8px 0; border-bottom:1px solid #eef0f7; }
</style>
""", unsafe_allow_html=True)


@st.cache_data(show_spinner=False)
def load(file_bytes, today):
    src = io.BytesIO(file_bytes) if file_bytes else DATA_DEFAULT
    return parse_workbook(src, today=today)


def kpi(col, label, value, sub=""):
    col.markdown(
        f'<div class="kpi"><div class="label">{label}</div>'
        f'<div class="value">{value}</div>'
        f'<div class="sub">{sub}</div></div>', unsafe_allow_html=True)


def pill(s, colors, textmap):
    bg = colors.get(s, "#E5E7EB")
    fg = textmap.get(s, "#374151")
    return f'<span class="pill" style="background:{bg};color:{fg}">{s}</span>'


def progress_bar(pct, status):
    color = COMP_COLORS.get(status, "#A7F3D0")
    return (f'<div class="bar-wrap"><div class="bar-fill" '
            f'style="width:{pct}%;background:{color}"></div></div>')


def style_cells(df, mapping, subset):
    def _fn(v):
        if v in mapping[0]:
            return f"background-color:{mapping[0][v]};color:{mapping[1][v]};font-weight:600;"
        return ""
    try:
        styler = df.style
        m = getattr(styler, "map", None) or styler.applymap
        return m(_fn, subset=subset)
    except Exception:  # noqa
        return df


def style_portfolio(df):
    """Tint 1st Year Testing columns (blue) and FAC columns (peach); keep the
    Status cells coloured by completion so both groups are easy to tell apart."""
    fyt_cols = ["FYT Status", "FYT %", "FYT Date"]
    fac_cols = ["FAC Status", "FAC %", "FAC Date"]

    def _row(row):
        out = []
        for col in df.columns:
            style = ""
            if col in fyt_cols:
                style = f"background-color:{FYT_COL_BG};"
            elif col in fac_cols:
                style = f"background-color:{FAC_COL_BG};"
            if col in ("FYT Status", "FAC Status"):
                v = row[col]
                if v in COMP_COLORS:
                    style = (f"background-color:{COMP_COLORS[v]};"
                             f"color:{COMP_TEXT[v]};font-weight:700;")
            out.append(style)
        return out

    try:
        return df.style.apply(_row, axis=1)
    except Exception:  # noqa
        return df


# ---- sidebar ----------------------------------------------------------------
with st.sidebar:
    st.markdown("### 🌤️ FAC Dashboard")
    st.caption("Ampyr Solar Europe · FAC & 1st Year Testing")
    up = st.file_uploader("Upload updated FAC Tracker (.xlsx)", type=["xlsx"])
    st.caption("The dashboard refreshes automatically with your latest file.")
    today = st.date_input("Reference date ('today')", value=dt.date(2026, 7, 21),
                          help="Date-timing labels are calculated relative to this date.")
    st.divider()

file_bytes = up.getvalue() if up else None

if not file_bytes and DATA_DEFAULT is None:
    st.title("FAC & 1st Year Testing Tracker")
    st.info("👋 No data file found yet. Use the **Upload updated FAC Tracker (.xlsx)** "
            "box in the sidebar, or add `data/FAC_Tracker_Gantt.xlsx` to the repo.")
    st.stop()

try:
    data = load(file_bytes, today)
except Exception as e:  # noqa
    st.error(f"Could not read the workbook: {e}")
    st.info("Try uploading the Excel file using the sidebar box on the left.")
    st.stop()

sites = data["sites"].copy()
ms = data["milestones"].copy()
tasks = data["tasks"].copy()
has_status = data.get("has_status", False)

with st.sidebar:
    st.markdown("#### Filters")
    countries = sorted([c for c in sites["Country"].dropna().unique() if c])
    contractors = sorted([c for c in sites["Contractor"].dropna().unique() if c])
    sel_country = st.multiselect("Country", countries, default=countries)
    sel_contractor = st.multiselect("O&M Contractor", contractors, default=contractors)
    if up:
        st.success("Using your uploaded file ✔")
    else:
        st.info("Showing the bundled tracker. Upload a file to update.")
    if not has_status:
        st.warning("No 'Status' column found — showing date timing only.")

mask = sites["Country"].isin(sel_country) & sites["Contractor"].isin(sel_contractor)
sites_f = sites[mask]
keep = set(sites_f["Site"])
ms_f = ms[ms["Site"].isin(keep)]
tasks_f = tasks[tasks["Site"].isin(keep)]

# ---- header -----------------------------------------------------------------
st.markdown("# FAC & 1st Year Testing Tracker")
st.caption(f"Reference date: **{today:%d %b %Y}** · {len(sites_f)} sites · "
           f"{len(tasks_f)} activities")

# ---- KPI row ----------------------------------------------------------------
n_done = int((tasks_f["Status"] == "Done").sum())
n_tasks = len(tasks_f)
pct_all = round(100 * n_done / n_tasks) if n_tasks else 0
fac_dates = sites_f["FAC Date"].dropna()
c1, c2, c3, c4, c5 = st.columns(5)
kpi(c1, "Sites tracked", len(sites_f), f"{sites_f['Country'].nunique()} countries")
kpi(c2, "FYT completed", int((sites_f["FYT Status"] == "Completed").sum()),
    f"of {len(sites_f)} sites")
kpi(c3, "FAC completed", int((sites_f["FAC Status"] == "Completed").sum()),
    f"of {len(sites_f)} sites")
kpi(c4, "Tasks done", f"{n_done}/{n_tasks}", f"{pct_all}% complete")
next_fac = fac_dates[fac_dates >= today].min() if (fac_dates >= today).any() else None
kpi(c5, "Next FAC date", f"{next_fac:%d %b %Y}" if next_fac else "—", "earliest upcoming")

st.write("")

tab1, tab2, tab3, tab4 = st.tabs(
    ["📋 Site Overview", "📅 FAC Status & Dates",
     "🧪 1st Year Performance", "🗂️ Activities & Gantt"])

# === TAB 1 : Site overview ===================================================
with tab1:
    left, right = st.columns([1.5, 1])
    with left:
        st.subheader("Portfolio")
        show = sites_f.copy()
        show["FAC Date"] = show["FAC Date"].apply(lambda d: f"{d:%d %b %Y}" if pd.notnull(d) else "TBC")
        show["FYT Date"] = show["FYT Date"].apply(lambda d: f"{d:%d %b %Y}" if pd.notnull(d) else "TBC")
        show["FYT %"] = show["FYT Progress %"].astype(int).astype(str) + "%"
        show["FAC %"] = show["FAC Progress %"].astype(int).astype(str) + "%"
        show = show[["Site", "Country", "Contractor",
                     "FYT Status", "FYT %", "FYT Date",
                     "FAC Status", "FAC %", "FAC Date"]]
        st.caption("🟦 blue = 1st Year Testing columns   🟧 peach = FAC columns")
        st.dataframe(style_portfolio(show),
                     use_container_width=True, hide_index=True, height=560)
    with right:
        st.subheader("First Year Testing status")
        vc = sites_f["FYT Status"].value_counts()
        fig = px.pie(values=vc.values, names=vc.index, hole=0.55,
                     color=vc.index, color_discrete_map=COMP_COLORS)
        fig.update_traces(textinfo="value")
        fig.update_layout(margin=dict(t=6, b=6, l=6, r=6), height=240,
                          legend=dict(orientation="h", y=-0.15),
                          paper_bgcolor="rgba(0,0,0,0)")
        st.plotly_chart(fig, use_container_width=True, key="fyt_pie")

        st.subheader("FAC status")
        vc2 = sites_f["FAC Status"].value_counts()
        fig2 = px.pie(values=vc2.values, names=vc2.index, hole=0.55,
                      color=vc2.index, color_discrete_map=COMP_COLORS)
        fig2.update_traces(textinfo="value")
        fig2.update_layout(margin=dict(t=6, b=6, l=6, r=6), height=240,
                           legend=dict(orientation="h", y=-0.15),
                           paper_bgcolor="rgba(0,0,0,0)")
        st.plotly_chart(fig2, use_container_width=True, key="fac_pie")

# === TAB 2 : FAC status & dates =============================================
with tab2:
    st.subheader("Milestone timeline (FYT → FAC)")
    tl = ms_f.dropna(subset=["Date"]).copy()
    if not tl.empty:
        piv = tl.pivot_table(index="Site", columns="Milestone", values="Date",
                             aggfunc="first").reset_index()
        rows = []
        for _, r in piv.iterrows():
            fyt_d, fac_d = r.get("First Year Testing"), r.get("FAC")
            if pd.notnull(fyt_d) and pd.notnull(fac_d):
                rows.append({"Site": r["Site"], "Start": fyt_d, "Finish": fac_d})
        if rows:
            gdf = pd.DataFrame(rows).sort_values("Start")
            figt = px.timeline(gdf, x_start="Start", x_end="Finish", y="Site",
                               color_discrete_sequence=["#BFD7FF"])
            figt.update_yaxes(autorange="reversed")
            figt.add_vline(x=pd.Timestamp(today), line_width=2,
                           line_dash="dash", line_color="#F9A8B4")
            fyt_rows = tl[tl.Milestone == "First Year Testing"]
            fac_rows = tl[tl.Milestone == "FAC"]
            fyt_clr = [GREEN_DONE if st_ == "Completed" else "#F9A8D4"
                       for st_ in fyt_rows["Status"]]
            fac_clr = [GREEN_DONE if st_ == "Completed" else "#93C5FD"
                       for st_ in fac_rows["Status"]]
            figt.add_trace(go.Scatter(
                x=fyt_rows["Date"], y=fyt_rows["Site"],
                mode="markers", name="FYT",
                text=fyt_rows["Status"], hovertemplate="%{y}<br>FYT: %{text}<extra></extra>",
                marker=dict(size=13, color=fyt_clr, line=dict(width=1, color="#fff"))))
            figt.add_trace(go.Scatter(
                x=fac_rows["Date"], y=fac_rows["Site"],
                mode="markers", name="FAC",
                text=fac_rows["Status"], hovertemplate="%{y}<br>FAC: %{text}<extra></extra>",
                marker=dict(size=14, color=fac_clr, symbol="diamond",
                            line=dict(width=1, color="#fff"))))
            # legend hint for the green = completed convention
            figt.add_trace(go.Scatter(
                x=[None], y=[None], mode="markers", name="Completed",
                marker=dict(size=13, color=GREEN_DONE, line=dict(width=1, color="#fff"))))
            figt.update_layout(height=460, margin=dict(t=20, b=10, l=10, r=10),
                               paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                               legend=dict(orientation="h", y=1.08))
            st.plotly_chart(figt, use_container_width=True, key="timeline")
            st.caption("Bar spans First Year Testing start → Final Acceptance. "
                       "Dots turn green when that milestone is Completed. "
                       "Pink dashed line = reference date.")

    colA, colB = st.columns(2)
    with colA:
        st.subheader("Upcoming milestone dates")
        up_ms = ms_f[(ms_f["Date"].notna()) & (ms_f["Date"] >= today)].sort_values("Date")
        if up_ms.empty:
            st.info("No upcoming milestone dates after the reference date.")
        else:
            for _, r in up_ms.head(12).iterrows():
                days = (r["Date"] - today).days
                st.markdown(
                    f"**{r['Site']}** — {r['Milestone']}  "
                    f"{pill(r['Status'], COMP_COLORS, COMP_TEXT)} "
                    f"{pill(r['Timing'], TIME_COLORS, TIME_TEXT)}<br>"
                    f"<span style='color:#8a889f'>{r['Date']:%d %b %Y} · in {days} days · "
                    f"{r['Progress %']}% done</span>", unsafe_allow_html=True)
    with colB:
        st.subheader("Milestones not yet complete")
        inc = ms_f[ms_f["Status"] != "Completed"].sort_values("Date", na_position="last")
        if inc.empty:
            st.success("All milestones complete 🎉")
        else:
            for _, r in inc.head(14).iterrows():
                dtxt = f"{r['Date']:%d %b %Y}" if pd.notnull(r["Date"]) else "date TBC"
                st.markdown(
                    f"**{r['Site']}** — {r['Milestone']}  "
                    f"{pill(r['Status'], COMP_COLORS, COMP_TEXT)}<br>"
                    f"<span style='color:#8a889f'>{dtxt} · {r['Progress %']}% done "
                    f"({r['Done']}/{r['Total']} tasks)</span>", unsafe_allow_html=True)

# === TAB 3 : 1st year performance ===========================================
with tab3:
    st.subheader("First Year Testing — completion")
    fy = sites_f.copy()
    c1, c2, c3, c4 = st.columns(4)
    kpi(c1, "Completed", int((fy["FYT Status"] == "Completed").sum()), "sites")
    kpi(c2, "In progress", int((fy["FYT Status"] == "In progress").sum()), "sites")
    kpi(c3, "Not started", int((fy["FYT Status"] == "Not started").sum()), "sites")
    fyt_tasks = tasks_f[tasks_f["Phase"] == "First Year Testing"]
    fd = int((fyt_tasks["Status"] == "Done").sum())
    kpi(c4, "FYT tasks done", f"{fd}/{len(fyt_tasks)}",
        f"{round(100*fd/len(fyt_tasks)) if len(fyt_tasks) else 0}% complete")
    st.write("")

    st.markdown("**Progress by site**")
    for _, r in fy.sort_values("FYT Progress %", ascending=False).iterrows():
        pctv = int(r["FYT Progress %"])
        cL, cM, cR = st.columns([2, 5, 2])
        cL.markdown(f"**{r['Site']}**")
        cM.markdown(progress_bar(pctv, r["FYT Status"]), unsafe_allow_html=True)
        cR.markdown(f"{pctv}%  {pill(r['FYT Status'], COMP_COLORS, COMP_TEXT)}",
                    unsafe_allow_html=True)

    st.write("")
    l, r = st.columns(2)
    with l:
        st.markdown("**FYT status breakdown**")
        vc = fy["FYT Status"].value_counts()
        figp = px.pie(values=vc.values, names=vc.index, hole=0.55,
                      color=vc.index, color_discrete_map=COMP_COLORS)
        figp.update_traces(textinfo="value")
        figp.update_layout(height=300, margin=dict(t=10, b=10, l=10, r=10),
                           legend=dict(orientation="h", y=-0.1),
                           paper_bgcolor="rgba(0,0,0,0)")
        st.plotly_chart(figp, use_container_width=True, key="fyt_breakdown")
    with r:
        st.markdown("**FYT tasks: done vs pending by category**")
        ft = fyt_tasks.copy()
        ft = ft[ft["Category"] != ""]
        if not ft.empty:
            g = ft.groupby(["Category", "Status"]).size().reset_index(name="n")
            figb = px.bar(g, x="Category", y="n", color="Status",
                          color_discrete_map=TASK_COLORS, barmode="stack")
            figb.update_layout(height=300, margin=dict(t=10, b=10, l=10, r=10),
                               xaxis_title="", yaxis_title="tasks",
                               legend=dict(orientation="h", y=1.1),
                               paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)")
            st.plotly_chart(figb, use_container_width=True, key="cat_bar")

# === TAB 4 : activities & gantt =============================================
with tab4:
    st.subheader("Activity Gantt")
    chips = "".join(f'<span class="legend-chip" style="background:{c}">{k}</span>'
                    for k, c in PASTEL.items())
    st.markdown(chips, unsafe_allow_html=True)

    site_opts = list(sites_f["Site"])
    sel_site = st.selectbox("Select a site", site_opts) if site_opts else None
    if sel_site:
        gt = tasks_f[(tasks_f["Site"] == sel_site) &
                     tasks_f["Start"].notna() & tasks_f["End"].notna()].copy()
        if gt.empty:
            st.info("No scheduled Gantt bars for this site (dates To Be Confirmed).")
        else:
            gt["Label"] = gt["Sub-Task"].where(gt["Sub-Task"] != "", gt["Task"])
            gt["Label"] = gt.apply(
                lambda x: ("✓ " if x["Status"] == "Done" else "") + x["Label"], axis=1)
            gt = gt.iloc[::-1]
            figg = px.timeline(gt, x_start="Start", x_end="End", y="Label",
                               color="Category", color_discrete_map=PASTEL,
                               hover_data=["Phase", "Task", "Status",
                                           "Responsibility", "Accepted By"])
            figg.add_vline(x=pd.Timestamp(today), line_width=2,
                           line_dash="dash", line_color="#F9A8B4")
            figg.update_layout(height=max(400, 22 * len(gt)),
                               margin=dict(t=20, b=10, l=10, r=10),
                               legend=dict(orientation="h", y=1.04),
                               paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                               yaxis_title="")
            st.plotly_chart(figg, use_container_width=True, key="gantt")
            done_here = int((gt["Status"] == "Done").sum())
            st.caption(f"✓ marks completed tasks · {done_here}/{len(gt)} done for this site.")

    st.subheader("Completed tasks")
    done_all = tasks_f[tasks_f["Status"] == "Done"].copy()
    done_all["End"] = done_all["End"].apply(
        lambda d: f"{d:%d %b %Y}" if pd.notnull(d) else "")
    fyt_done = done_all[done_all["Phase"] == "First Year Testing"]
    fac_done = done_all[done_all["Phase"] == "FAC"]
    dcol1, dcol2 = st.columns(2)
    with dcol1:
        st.markdown(f"**🧪 1st Year Testing — {len(fyt_done)} tasks done**")
        if fyt_done.empty:
            st.info("No 1st Year Testing tasks marked done yet.")
        else:
            st.dataframe(fyt_done[["Site", "Task", "Sub-Task", "Accepted By", "End"]],
                         use_container_width=True, hide_index=True, height=320)
    with dcol2:
        st.markdown(f"**📜 FAC — {len(fac_done)} tasks done**")
        if fac_done.empty:
            st.info("No FAC tasks marked done yet.")
        else:
            st.dataframe(fac_done[["Site", "Task", "Sub-Task", "Accepted By", "End"]],
                         use_container_width=True, hide_index=True, height=320)
    st.write("")

    st.subheader("All activities")
    tbl = tasks_f.copy()
    for c in ["Start", "End"]:
        tbl[c] = tbl[c].apply(lambda d: f"{d:%d %b %Y}" if pd.notnull(d) else "")
    f1, f2 = st.columns(2)
    fcat = f1.multiselect("Filter by category",
                          sorted([c for c in tbl["Category"].unique() if c]))
    fstat = f2.multiselect("Filter by status", ["Done", "Pending"])
    if fcat:
        tbl = tbl[tbl["Category"].isin(fcat)]
    if fstat:
        tbl = tbl[tbl["Status"].isin(fstat)]
    view = tbl[["Site", "Phase", "Task", "Sub-Task", "Category", "Status",
                "Responsibility", "Accepted By", "Start", "End"]]
    st.dataframe(style_cells(view, (TASK_COLORS, TASK_TEXT), ["Status"]),
                 use_container_width=True, hide_index=True, height=460)
    st.download_button("⬇️ Download filtered activities (CSV)",
                       view.to_csv(index=False).encode(),
                       file_name="fac_activities.csv", mime="text/csv")

st.divider()
st.caption("Headline status comes from the workbook's Status column "
           "(Completed / In progress / Not started, based on tasks marked 'Done'). "
           "Date timing (Date passed / Due soon / Scheduled) is relative to the "
           "reference date and shown as a secondary indicator.")
