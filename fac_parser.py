"""
Parser for the Ampyr FAC Tracker (Gantt) Excel workbook.

Layout (auto-detected, so it survives inserted columns):
  * Columns A-H  : Site Name, FAC Date, O&M Contractor, Task, Sub-Task,
                   Responsibility, Accepted By, Dates as per testing.
  * A "Status"   : optional column (value "Done" = task completed, blank = not).
  * Gantt columns: one week each; coloured cells are task bars, the fill colour
                   encodes the category. The first Gantt column is found from the
                   4-digit year in row 1.
  * A pink "today marker" cell anchors the weekly columns to real dates
    (marker week = 14-Jul-2026, 7 days per column).
"""

from __future__ import annotations

from datetime import date, datetime, timedelta

import openpyxl
import pandas as pd

CATEGORY_COLORS = {
    "FFC4B5FD": "Prerequisite",
    "FF6EE7B7": "Compliance",
    "FFFDE68A": "Closure",
    "FFFED7AA": "Financial",
    "FFA5F3FC": "Technical",
    "FFFBCFE8": "Testing",
    "FFBFDBFE": "Document",
}

TODAY_MARKER_FILL = "FFFEE2E2"
DEFAULT_ANCHOR_DATE = date(2026, 7, 14)
FIRST_METADATA_ROW = 4

# metadata column positions (1-based) — stable, Status/Gantt are added after these
COL_SITE, COL_FACDATE, COL_OM = 1, 2, 3
COL_TASK, COL_SUBTASK, COL_RESP, COL_ACCEPT, COL_TESTDATE = 4, 5, 6, 7, 8

DONE_TOKENS = {"done", "yes", "y", "complete", "completed", "✓", "true", "x"}


def _fill_hex(cell):
    fill = cell.fill
    if fill and fill.patternType:
        rgb = fill.fgColor.rgb
        if isinstance(rgb, str):
            return rgb
    return None


def _detect_layout(ws):
    """Return (status_col or None, first_gantt_col)."""
    status_col = None
    gantt_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str) and v.strip().lower() == "status":
            status_col = c
        if gantt_col is None and v is not None:
            s = str(v).strip()
            if s.isdigit() and len(s) == 4 and 2000 <= int(s) <= 2100:
                gantt_col = c
    if gantt_col is None:
        gantt_col = (status_col + 1) if status_col else 9
    return status_col, gantt_col


def _find_anchor(ws, first_gantt_col):
    for r in range(1, 4):
        for c in range(first_gantt_col, ws.max_column + 1):
            if _fill_hex(ws.cell(row=r, column=c)) == TODAY_MARKER_FILL:
                return c, DEFAULT_ANCHOR_DATE
    return first_gantt_col + 80, DEFAULT_ANCHOR_DATE


def _col_to_date(col, anchor_col, anchor_date):
    return anchor_date + timedelta(weeks=(col - anchor_col))


def _is_done(v):
    return isinstance(v, str) and v.strip().lower() in DONE_TOKENS or v is True


def _parse_header(text):
    parts = [p.strip() for p in str(text).split("|")]
    name = parts[0].strip() if parts else str(text).strip()
    country, fac_raw, contractor = "", "", ""
    for p in parts[1:]:
        low = p.lower()
        if low.startswith("fac"):
            fac_raw = p.split(":", 1)[1].strip() if ":" in p else ""
        elif low.startswith("o&m") or low.startswith("o & m"):
            contractor = p.split(":", 1)[1].strip() if ":" in p else ""
        elif p and not country:
            country = p
    return name, country, contractor, fac_raw


def _parse_date_any(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    s = str(v).strip()
    if not s or s in {"—", "-", "TBC", "TBD"}:
        return None
    for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _timing(d, today):
    if d is None:
        return "TBC"
    delta = (d - today).days
    if delta < 0:
        return "Date passed"
    if delta <= 90:
        return "Due soon"
    return "Scheduled"


def _completion_status(done, total):
    if total == 0:
        return "No tasks"
    if done >= total:
        return "Completed"
    if done > 0:
        return "In progress"
    return "Not started"


def parse_workbook(source, today=None):
    if today is None:
        today = date.today()
    if isinstance(today, datetime):
        today = today.date()

    wb = openpyxl.load_workbook(source, data_only=True)
    ws = wb[wb.sheetnames[0]]
    status_col, first_gantt = _detect_layout(ws)
    anchor_col, anchor_date = _find_anchor(ws, first_gantt)

    sites, milestones, tasks = [], [], []
    cur_site = cur_phase = cur_task = None
    phase_agg = {}  # (site, phase) -> [done, total]

    for r in range(FIRST_METADATA_ROW, ws.max_row + 1):
        A = ws.cell(row=r, column=COL_SITE).value
        D = ws.cell(row=r, column=COL_TASK).value
        E = ws.cell(row=r, column=COL_SUBTASK).value
        F = ws.cell(row=r, column=COL_RESP).value
        G = ws.cell(row=r, column=COL_ACCEPT).value
        H = ws.cell(row=r, column=COL_TESTDATE).value
        S = ws.cell(row=r, column=status_col).value if status_col else None

        a_txt = str(A).strip() if A is not None else ""
        d_txt = str(D).strip() if D is not None else ""
        e_txt = str(E).strip() if E is not None else ""
        done_flag = _is_done(S)

        if a_txt and "|" in a_txt:
            name, country, contractor, fac_raw = _parse_header(a_txt)
            cur_site = {"Site": name, "Country": country, "Contractor": contractor,
                        "FAC Date": _parse_date_any(fac_raw), "FYT Date": None}
            sites.append(cur_site)
            cur_phase = cur_task = None
            continue

        if cur_site is None:
            continue
        if a_txt and a_txt.lower() == cur_site["Site"].lower():
            continue

        if d_txt.startswith("▸") or d_txt.startswith(">"):
            up = d_txt.upper()
            if "FIRST YEAR" in up:
                cur_phase = "First Year Testing"
            elif "FINAL ACCEPTANCE" in up:
                cur_phase = "FAC"
            else:
                cur_phase = d_txt.lstrip("▸> ").strip()
            phase_date = _parse_date_any(H)
            if cur_phase == "First Year Testing":
                cur_site["FYT Date"] = phase_date
            milestones.append({"Site": cur_site["Site"], "Country": cur_site["Country"],
                               "Contractor": cur_site["Contractor"],
                               "Milestone": cur_phase, "Date": phase_date,
                               "PhaseHeaderDone": done_flag})
            phase_agg.setdefault((cur_site["Site"], cur_phase), [0, 0])
            cur_task = None
            continue

        if d_txt:
            cur_task = d_txt
        if not (d_txt or e_txt):
            continue

        start_col = end_col = None
        counts = {}
        for c in range(first_gantt, ws.max_column + 1):
            hexc = _fill_hex(ws.cell(row=r, column=c))
            if hexc in CATEGORY_COLORS:
                if start_col is None:
                    start_col = c
                end_col = c
                cat = CATEGORY_COLORS[hexc]
                counts[cat] = counts.get(cat, 0) + 1
        category = max(counts, key=counts.get) if counts else ""
        start_date = _col_to_date(start_col, anchor_col, anchor_date) if start_col else None
        end_date = (_col_to_date(end_col, anchor_col, anchor_date) + timedelta(days=6)) if end_col else None

        key = (cur_site["Site"], cur_phase)
        agg = phase_agg.setdefault(key, [0, 0])
        agg[1] += 1
        if done_flag:
            agg[0] += 1

        tasks.append({"Site": cur_site["Site"], "Country": cur_site["Country"],
                      "Contractor": cur_site["Contractor"], "Phase": cur_phase or "",
                      "Task": cur_task or "", "Sub-Task": e_txt,
                      "Responsibility": str(F).strip() if F else "",
                      "Accepted By": str(G).strip() if G else "",
                      "Category": category, "Start": start_date, "End": end_date,
                      "Status": "Done" if done_flag else "Pending"})

    df_sites = pd.DataFrame(sites)
    df_ms = pd.DataFrame(milestones)
    df_tasks = pd.DataFrame(tasks)

    # ---- roll up completion onto milestones & sites ----
    def phase_done_total(site, phase):
        return phase_agg.get((site, phase), [0, 0])

    if not df_ms.empty:
        st_list, done_list, tot_list, pct_list, timing_list = [], [], [], [], []
        for _, r in df_ms.iterrows():
            d, t = phase_done_total(r["Site"], r["Milestone"])
            if t > 0 and d == t:
                header_done = True
            else:
                header_done = bool(r.get("PhaseHeaderDone"))
                if header_done:  # phase flagged done at header -> treat as complete
                    d, t = (t, t) if t else (1, 1)
            status = _completion_status(d, t)
            done_list.append(d); tot_list.append(t)
            pct_list.append(round(100 * d / t) if t else 0)
            st_list.append(status)
            timing_list.append(_timing(r["Date"], today))
        df_ms["Done"] = done_list
        df_ms["Total"] = tot_list
        df_ms["Progress %"] = pct_list
        df_ms["Status"] = st_list
        df_ms["Timing"] = timing_list

    if not df_sites.empty:
        for lab, phase in [("FYT", "First Year Testing"), ("FAC", "FAC")]:
            dts = df_sites["Site"].apply(lambda s: phase_done_total(s, phase))
            df_sites[f"{lab} Done"] = dts.apply(lambda x: x[0])
            df_sites[f"{lab} Total"] = dts.apply(lambda x: x[1])
            df_sites[f"{lab} Progress %"] = dts.apply(
                lambda x: round(100 * x[0] / x[1]) if x[1] else 0)
            df_sites[f"{lab} Status"] = dts.apply(lambda x: _completion_status(x[0], x[1]))
        # honour a phase-header "Done" flag even when task rows are blank
        if not df_ms.empty:
            for lab, phase in [("FYT", "First Year Testing"), ("FAC", "FAC")]:
                hdr = df_ms[df_ms["Milestone"] == phase].set_index("Site")["Status"]
                df_sites[f"{lab} Status"] = df_sites.apply(
                    lambda row: hdr.get(row["Site"], row[f"{lab} Status"]), axis=1)
        df_sites["FAC Timing"] = df_sites["FAC Date"].apply(lambda d: _timing(d, today))
        df_sites["FYT Timing"] = df_sites["FYT Date"].apply(lambda d: _timing(d, today))

    return {"sites": df_sites, "milestones": df_ms, "tasks": df_tasks,
            "today": today, "anchor_col": anchor_col, "anchor_date": anchor_date,
            "has_status": status_col is not None}


if __name__ == "__main__":
    import sys
    data = parse_workbook(sys.argv[1], today=date(2026, 7, 21))
    print("has_status:", data["has_status"])
    cols = ["Site", "FYT Done", "FYT Total", "FYT Status",
            "FAC Done", "FAC Total", "FAC Status"]
    print(data["sites"][cols].to_string(index=False))
    print("\nTask status counts:\n", data["tasks"]["Status"].value_counts())
